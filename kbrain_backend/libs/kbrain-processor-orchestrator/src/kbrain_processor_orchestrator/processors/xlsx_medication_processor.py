"""XLSX Medication processor for spreadsheets tagged with 'xlsx_z_lekami'."""

import os

import httpx
from loguru import logger

from kbrain_processor_orchestrator.base import BaseProcessor
from kbrain_processor_orchestrator.models import (
    DocumentInfo,
    ProcessingResult,
    QueueMessage,
)

TAG_NAME = "xlsx_z_lekami"


class XlsxMedicationProcessor(BaseProcessor):
    """
    Processor for XLSX files tagged with 'xlsx_z_lekami'.

    Applies alternative processing logic for medication spreadsheets.
    Only activates when the document has the matching tag.
    Sends extracted data to an external medication service (not RAGFlow).
    """

    SUPPORTED_EXTENSIONS = {"xlsx"}

    def __init__(
        self,
        service_url: str | None = None,
    ) -> None:
        self.service_url = service_url or os.getenv("XLSX_MEDICATION_SERVICE_URL")

        if self.service_url:
            logger.info(f"XLSX medication service URL: {self.service_url}")
        else:
            logger.warning(
                "No XLSX_MEDICATION_SERVICE_URL configured. "
                "Set XLSX_MEDICATION_SERVICE_URL env var or pass service_url."
            )

    @property
    def name(self) -> str:
        return "xlsx_medication_processor"

    async def can_process(self, document: DocumentInfo) -> bool:
        """Only process XLSX files that have the 'xlsx_z_lekami' tag."""
        if document.file_extension.lower() not in self.SUPPORTED_EXTENSIONS:
            return False

        has_tag = any(tag.name == TAG_NAME for tag in document.tags)
        if has_tag:
            logger.info(
                f"Document {document.original_name} has '{TAG_NAME}' tag "
                f"- will use medication processor"
            )
        return has_tag

    async def process(self, document: DocumentInfo) -> ProcessingResult:
        """
        Process an XLSX medication spreadsheet.

        Reads the XLSX, extracts rows, and sends to the medication service.
        """
        logger.info(
            f"Processing medication XLSX: {document.original_name} "
            f"(tag: {TAG_NAME})"
        )

        try:
            file_bytes = await self._download_file(document.download_url)

            if not file_bytes:
                return ProcessingResult(
                    success=False,
                    error_message="Failed to download XLSX file",
                )

            text_content = self._extract_medication_data(file_bytes, document.original_name)

            result_metadata: dict[str, object] = {
                "processor": self.name,
                "file_type": "xlsx",
                "tag": TAG_NAME,
                "file_size": len(file_bytes),
                "char_count": len(text_content),
                "word_count": len(text_content.split()),
                "text_content": text_content,
            }

            logger.info(
                f"Extracted {len(text_content)} chars from medication XLSX: "
                f"{document.original_name}"
            )

            # Send to medication service if configured
            if self.service_url:
                service_result = await self._send_to_service(
                    document, text_content, file_bytes
                )
                result_metadata["service"] = service_result

            return ProcessingResult(
                success=True,
                metadata=result_metadata,
            )

        except Exception as e:
            logger.error(
                f"Error processing medication XLSX {document.original_name}: {e}"
            )
            return ProcessingResult(
                success=False,
                error_message=str(e),
            )

    def _extract_medication_data(self, file_bytes: bytes, filename: str) -> str:
        """
        Extract medication data from XLSX bytes.

        Reads all sheets and converts each row into a readable text block.
        """
        import io
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            headers = [str(cell) if cell is not None else "" for cell in rows[0]]

            parts.append(f"## {sheet_name}\n")

            for row in rows[1:]:
                row_parts: list[str] = []
                for header, cell in zip(headers, row):
                    if cell is not None and str(cell).strip():
                        row_parts.append(f"{header}: {cell}")
                if row_parts:
                    parts.append(" | ".join(row_parts))

        wb.close()

        text_content = "\n".join(parts)
        logger.info(
            f"Extracted medication data from {filename}: "
            f"{len(parts)} lines from {len(wb.sheetnames)} sheet(s)"
        )
        return text_content

    async def _send_to_service(
        self,
        document: DocumentInfo,
        text_content: str,
        file_bytes: bytes,
    ) -> dict[str, object]:
        """
        Send extracted medication data to the external service.

        Args:
            document: Document info
            text_content: Extracted text content
            file_bytes: Raw XLSX file bytes

        Returns:
            Service response dict
        """
        if not self.service_url:
            return {"error": "Service URL not configured"}

        logger.info(
            f"Sending medication data to service: {self.service_url} "
            f"(file: {document.original_name})"
        )

        try:
            async with httpx.AsyncClient(timeout=120.0) as client:
                response = await client.post(
                    self.service_url,
                    files={"file": (document.original_name, file_bytes)},
                    data={
                        "document_id": str(document.id),
                        "scope_id": str(document.scope_id),
                        "original_name": document.original_name,
                        "text_content": text_content,
                    },
                )
                response.raise_for_status()

                result = response.json() if response.headers.get("content-type", "").startswith("application/json") else {}
                logger.info(
                    f"Medication service response: {response.status_code}"
                )
                return {"success": True, "status_code": response.status_code, **result}

        except Exception as e:
            logger.error(f"Failed to send to medication service: {e}")
            return {"success": False, "error": str(e)}

    async def _download_file(self, url: str | None) -> bytes | None:
        if not url:
            return None
        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                response = await client.get(url)
                response.raise_for_status()
                return response.content
        except Exception as e:
            logger.error(f"Failed to download file from {url}: {e}")
            return None

    async def on_delete(self, message: QueueMessage) -> ProcessingResult:
        logger.info(
            f"Cleanup for deleted medication XLSX {message.document_id} "
            f"(file: {message.original_name or 'unknown'})"
        )

        return ProcessingResult(
            success=True,
            metadata={
                "action": "delete",
                "document_id": str(message.document_id),
                "file_extension": message.file_extension,
            },
        )
